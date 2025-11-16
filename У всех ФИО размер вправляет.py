from openpyxl import load_workbook
def simple_fio_converter():
    # Открываем файл
    wb = load_workbook('1.xlsx')
    ws = wb['1']    
    # Обрабатываем каждую строку в столбце B
    for row in range(2, ws.max_row + 1):
        fio_cell = ws[f'B{row}']
        if fio_cell.value:
            # Преобразуем ФИО
            original = fio_cell.value
            normalized = ' '.join([word.capitalize() for word in str(original).split()])            
            # Записываем в столбец E
            ws[f'F{row}'] = normalized            
            print(f"Обработано: {original} -> {normalized}")    
    # Сохраняем результат
    wb.save('1_1.xlsx')
    print("Готово! Файл сохранен как '1_1.xlsx'")
# Запуск
simple_fio_converter()